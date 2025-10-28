import pandas as pd
import openpyxl
from datetime import datetime
import json

# Read the Excel file
file_path = r"C:\Users\blkw\OneDrive\Documents\Claude Code\Source 4 Industries\Bulk Bollards\S4 Bollards - First Order Data (Final).xlsx"

# Load the workbook to get sheet names
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
print("\n" + "="*80 + "\n")

# Read all sheets
sheets = {}
for sheet_name in wb.sheetnames:
    print(f"Reading sheet: {sheet_name}")
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    sheets[sheet_name] = df
    print(f"  Shape: {df.shape} (rows, columns)")
    print(f"  Columns: {list(df.columns)}")
    print(f"  First few rows:")
    print(df.head(3))
    print("\n" + "-"*80 + "\n")

# Save each sheet as CSV for easier analysis
for sheet_name, df in sheets.items():
    safe_name = sheet_name.replace(" ", "_").replace("/", "_")
    csv_path = f"C:\\Users\\blkw\\OneDrive\\Documents\\Claude Code\\Source 4 Industries\\Bulk Bollards\\{safe_name}.csv"
    df.to_csv(csv_path, index=False)
    print(f"Saved: {safe_name}.csv")

print("\n" + "="*80)
print("Analysis complete! All sheets saved as CSV files.")
