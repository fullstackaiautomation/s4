import pandas as pd
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load configuration
with open('config.json', 'r') as f:
    config = json.load(f)

month = config['month']
output_dir = config['paths']['output_dir'].replace("{month}", month)

# Load the CSV files
upload = pd.read_csv(os.path.join(output_dir, f"{month} Product Spend Upload.csv"))

# Define vendor category structure
vendor_categories = {
    'S4 Bollards': ['Bollard Covers', 'Crash Rated Bollards', 'Fixed Bollards', 'Flexible Bollards', 'Removable Bollards', 'Retractable Bollards'],
    'Handle-It': ['Floor Mounted Barrier', 'Forklift Wheel Stops', 'Guard Rail', 'Rack Protection', 'Stretch Wrap Machines'],
    'Casters': ['Bellman Casters', 'Gate Casters', 'General Casters', 'Heavy Duty / Container', 'High Temp Casters', 'Leveling Casters'],
    'Lincoln Industrial': ['Air Motors', 'Hoses', 'Kits', 'Other', 'Pumps', 'Quicklub'],
    'Noblelift': ['Battery, Charger, Accessories', 'Bigger Electric Equipment', 'EDGE Powered', 'Electric Pallet Jacks', 'Manual Pallet Jacks', 'Scissor Lifts', 'Straddle Stackers'],
    'B&P Manufacturing': ['Aristocrat', 'Carts', 'Dock Plates', 'Hand Truck Accessories', 'Hand Trucks', 'Ramps'],
    'Dutro': ['Accessories', 'Carts', 'Dollies', 'Hand Trucks', 'Mattress Moving Carts', 'Vending Machine Trucks'],
    'Reliance Foundry': ['Bollard Covers', 'Crash Rated Bollards', 'Decorative Bollards', 'Fixed Bollards', 'Flexible Bollards', 'Fold Down Bollards', 'Removable Bollards', 'Retractable Bollards'],
    'Ekko Lifts': ['Electric Forklifts', 'Electric Pallet Jacks', 'Electric Straddle Stackers', 'Electric Walkie Stackers', 'Manual Pallet Jacks', 'Other'],
    "Adrian's Safety": ['Cargo Safety', 'Pallet Rack Safety Straps', 'Pallet Rack Safety Netting'],
    'Sentry Protection': ['ST - Accessories', 'ST - Collision Sentry', 'ST - Column Protectors'],
    'Little Giant': ['Cabinet', 'Carts', 'Dollies', 'Gas Cylinder', 'Rack', 'Tables'],
    'Merrick Machine': ['Accessories', 'Auto Dollies', 'Auto Rotisseries', 'Flat Top Dollies', 'Lifts, Rack, Stands', 'Other Dollies'],
    'Wesco': ['Accessories & Other', 'Carts, Hand Trucks & Dollies', 'Dock Equipment', 'Drum Equipment', 'Lifts & Stackers', 'Pallet Jacks'],
    'Valley Craft': ['Accessories', 'Cabinets & Desks', 'Carts', 'Dumpers & Lifts', 'Hand Trucks & Dollies', 'Other'],
    'Bluff Manufacturing': ['Bollards & Protectors', 'Dock Boards', 'Edge of Dock Levelers', 'Other', 'Ramps', 'Stairways'],
    'Meco-Omaha': ['Cantilever', 'Carts & Dollies', 'Guard Rail', 'Hoppers', 'Other', 'Racking'],
    'Apollo Forklift': ['AF - Electric Pallet Jacks', 'AF - Electric Stackers', 'AF - Manual Pallet Jacks', 'AF - Manual Stackers', 'AF - Order Pickers', 'AF - Scissor Lifts'],
}

# Main vendors list
main_vendors_list = list(vendor_categories.keys())

print("Building vendor category breakdown...")
print(f"Main vendors: {len(main_vendors_list)}")

# Define caster vendors early (must match exact vendor names in data)
caster_vendors = ['Caster Depot', 'Dh International', 'Durable Superior Casters']

# Calculate spend by vendor and category
vendor_spend = {}
category_spend = {}  # Will store {vendor: {category: spend}}

for vendor in upload['Vendor'].unique():
    vendor_data = upload[upload['Vendor'] == vendor]
    spend = sum([float(x.replace('$','').replace(',','')) for x in vendor_data['Ad Spend']])
    vendor_spend[vendor] = spend

    # Calculate category spend for this vendor
    category_spend[vendor] = {}
    for category in vendor_data['Product Category'].unique():
        if pd.notna(category) and category != '' and str(category).upper() != 'BLANK':
            cat_data = vendor_data[vendor_data['Product Category'] == category]
            cat_spend = sum([float(x.replace('$','').replace(',','')) for x in cat_data['Ad Spend']])
            category_spend[vendor][category] = cat_spend

# Identify "All Other Vendors" (excluding the caster component vendors)
main_vendor_names = set(vendor_categories.keys())
other_vendors = [v for v in vendor_spend.keys() if v not in main_vendor_names and v not in caster_vendors]
# Sort by spend descending
other_vendors = sorted(other_vendors, key=lambda v: vendor_spend[v], reverse=True)
other_vendors_spend = sum([vendor_spend[v] for v in other_vendors])

print(f"Other vendors: {len(other_vendors)}")
print(f"Other vendors spend: ${other_vendors_spend:,.2f}")

# Create Excel file
output_file = os.path.join(output_dir, f"{month} Product Spend Report.xlsx")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
vendor_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
vendor_font = Font(bold=True, size=11)
category_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
category_font = Font(size=10)
other_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
other_font = Font(bold=True, color="C65911")
center_align = Alignment(horizontal="center", vertical="center")
currency_format = '$#,##0.00'

# Border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: Upload Sheet
print("Creating Sheet 1: Product Spend Upload...")
ws1 = wb.create_sheet("Product Spend Upload")
upload_df = pd.read_csv("2025-10 Product Spend Upload.csv")

# Add headers
for c_idx, col in enumerate(upload_df.columns, 1):
    cell = ws1.cell(row=1, column=c_idx, value=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

# Add data rows (no alternating colors)
for r_idx, row in enumerate(upload_df.values, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx+1, column=c_idx, value=value)
        cell.border = thin_border

# Set column widths (narrower)
for i in range(1, len(upload_df.columns) + 1):
    ws1.column_dimensions[get_column_letter(i)].width = 14

# Sheet 2: Missing Categories
print("Creating Sheet 2: Missing Categories...")
ws2 = wb.create_sheet("Missing Categories")
missing_cats = pd.read_csv(f"{month} Missing Product Categories.csv")

# Add headers
for c_idx, col in enumerate(missing_cats.columns, 1):
    cell = ws2.cell(row=1, column=c_idx, value=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

# Add data rows
for r_idx, row in enumerate(missing_cats.values, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx+1, column=c_idx, value=value)
        cell.border = thin_border

# Set narrower column widths
for i in range(1, len(missing_cats.columns) + 1):
    ws2.column_dimensions[get_column_letter(i)].width = 14

# Sheet 3: Vendor Category Breakdown
print("Creating Sheet 3: Vendor Breakdown...")
ws3 = wb.create_sheet("Vendor Breakdown")

# Calculate total ad spend and revenue
total_ad_spend = sum(vendor_spend.values())
total_revenue = 0
for vendor_data in [upload[upload['Vendor'] == v] for v in upload['Vendor'].unique()]:
    revenue = sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                   for x in vendor_data['Revenue']])
    total_revenue += revenue

# Handle Casters (sum of Caster Depot, DH International, Durable Superior Casters)
# Check what the actual vendor names are
caster_vendors = [v for v in vendor_spend.keys() if v in ['Caster Depot', 'Durable Superior Casters'] or 'international' in v.lower()]
casters_total = sum([vendor_spend.get(v, 0) for v in caster_vendors])

row = 1
# Add summary row
cell = ws3.cell(row=row, column=1, value="TOTAL")
cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
cell.font = Font(bold=True, color="FFFFFF", size=12)
cell.border = thin_border

cell = ws3.cell(row=row, column=2, value=f"${total_ad_spend:,.2f}")
cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
cell.font = Font(bold=True, color="FFFFFF", size=12)
cell.border = thin_border
cell.number_format = currency_format

cell = ws3.cell(row=row, column=3, value="")
cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
cell.border = thin_border

cell = ws3.cell(row=row, column=4, value=f"${total_revenue:,.2f}")
cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
cell.font = Font(bold=True, color="FFFFFF", size=12)
cell.border = thin_border
cell.number_format = currency_format

row = 2
# Add header row
cell = ws3.cell(row=row, column=1, value="")
cell.border = thin_border

cell = ws3.cell(row=row, column=2, value="Ad Spend")
cell.font = Font(bold=True, size=10)
cell.border = thin_border

cell = ws3.cell(row=row, column=3, value="")
cell.border = thin_border

cell = ws3.cell(row=row, column=4, value="Revenue")
cell.font = Font(bold=True, size=10)
cell.border = thin_border

row = 3  # Start vendor data after header row

for vendor_name in main_vendors_list:
    categories = vendor_categories[vendor_name]

    # Calculate vendor total
    if vendor_name == 'Casters':
        vendor_total = casters_total
    else:
        vendor_total = vendor_spend.get(vendor_name, 0)

    # Calculate vendor revenue
    if vendor_name == 'Casters':
        vendor_revenue = 0
        for cv in caster_vendors:
            cv_data = upload[upload['Vendor'] == cv]
            vendor_revenue += sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                                   for x in cv_data['Revenue']])
    else:
        vendor_data = upload[upload['Vendor'] == vendor_name]
        vendor_revenue = sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                              for x in vendor_data['Revenue']])

    # Vendor header
    cell = ws3.cell(row=row, column=1, value=f"{vendor_name}")
    cell.fill = vendor_fill
    cell.font = vendor_font
    cell.border = thin_border

    cell = ws3.cell(row=row, column=2, value=f"${vendor_total:,.2f}")
    cell.fill = vendor_fill
    cell.font = vendor_font
    cell.border = thin_border
    cell.number_format = currency_format

    # Empty column (C)
    cell = ws3.cell(row=row, column=3, value="")
    cell.fill = vendor_fill
    cell.border = thin_border

    # Revenue column
    cell = ws3.cell(row=row, column=4, value=f"${vendor_revenue:,.2f}")
    cell.fill = vendor_fill
    cell.font = vendor_font
    cell.border = thin_border
    cell.number_format = currency_format

    row += 1

    # Categories
    for category in categories:
        cell = ws3.cell(row=row, column=1, value=f"  {category}")
        cell.fill = category_fill
        cell.font = category_font
        cell.border = thin_border

        # Category total from actual data (if available)
        cat_total = 0.0
        cat_revenue = 0.0

        if vendor_name == 'Casters':
            # For Casters, sum across all three caster vendors
            for cv in caster_vendors:
                if cv in category_spend and category in category_spend[cv]:
                    cat_total += category_spend[cv][category]
                # Also sum revenue for this category
                cv_cat_data = upload[(upload['Vendor'] == cv) & (upload['Product Category'] == category)]
                cat_revenue += sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                                    for x in cv_cat_data['Revenue']])
        else:
            # For regular vendors, get from the vendor's category spend
            if vendor_name in category_spend and category in category_spend[vendor_name]:
                cat_total = category_spend[vendor_name][category]
            # Get revenue for this category
            cat_data = upload[(upload['Vendor'] == vendor_name) & (upload['Product Category'] == category)]
            cat_revenue = sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                               for x in cat_data['Revenue']])

        cell = ws3.cell(row=row, column=2, value=f"${cat_total:,.2f}" if cat_total > 0 else "")
        cell.border = thin_border
        if cat_total > 0:
            cell.number_format = currency_format

        # Empty column (C)
        cell = ws3.cell(row=row, column=3, value="")
        cell.fill = category_fill
        cell.border = thin_border

        # Category revenue
        cell = ws3.cell(row=row, column=4, value=f"${cat_revenue:,.2f}" if cat_revenue > 0 else "")
        cell.fill = category_fill
        cell.font = category_font
        cell.border = thin_border
        if cat_revenue > 0:
            cell.number_format = currency_format

        row += 1

    row += 1

# All Other Vendors section
# Calculate total revenue for all other vendors
all_other_vendors_revenue = 0
for vendor in other_vendors:
    vendor_data = upload[upload['Vendor'] == vendor]
    revenue = sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                   for x in vendor_data['Revenue']])
    all_other_vendors_revenue += revenue

cell = ws3.cell(row=row, column=1, value="All Other Vendors")
cell.fill = other_fill
cell.font = other_font
cell.border = thin_border

cell = ws3.cell(row=row, column=2, value=f"${other_vendors_spend:,.2f}")
cell.fill = other_fill
cell.font = other_font
cell.border = thin_border
cell.number_format = currency_format

cell = ws3.cell(row=row, column=3, value="")
cell.fill = other_fill
cell.border = thin_border

cell = ws3.cell(row=row, column=4, value=f"${all_other_vendors_revenue:,.2f}")
cell.fill = other_fill
cell.font = other_font
cell.border = thin_border
cell.number_format = currency_format

row += 1

for vendor in other_vendors:  # Already sorted by spend descending
    spend = vendor_spend[vendor]

    # Calculate revenue for this vendor
    vendor_data = upload[upload['Vendor'] == vendor]
    revenue = sum([float(x.replace('$','').replace(',','')) if x and isinstance(x, str) and x.strip() else 0
                   for x in vendor_data['Revenue']])

    cell = ws3.cell(row=row, column=1, value=f"  - {vendor}")
    cell.fill = other_fill
    cell.font = other_font
    cell.border = thin_border

    cell = ws3.cell(row=row, column=2, value=f"${spend:,.2f}")
    cell.fill = other_fill
    cell.font = other_font
    cell.border = thin_border
    cell.number_format = currency_format

    # Empty column (C)
    cell = ws3.cell(row=row, column=3, value="")
    cell.fill = other_fill
    cell.border = thin_border

    cell = ws3.cell(row=row, column=4, value=f"${revenue:,.2f}")
    cell.fill = other_fill
    cell.font = other_font
    cell.border = thin_border
    cell.number_format = currency_format

    row += 1

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 4
ws3.column_dimensions['D'].width = 14

# Save with temporary name first
temp_file = os.path.join(output_dir, f"{month} Product Spend Report_TEMP.xlsx")
wb.save(temp_file)

# Try to replace the original file
import os
import shutil
if os.path.exists(output_file):
    try:
        os.remove(output_file)
    except:
        try:
            shutil.move(temp_file, output_file, copy_function=shutil.copy2)
            temp_file = None
        except:
            pass

if temp_file and os.path.exists(temp_file):
    shutil.move(temp_file, output_file, copy_function=shutil.copy2)

print(f"\nCreated: {output_file}")
print(f"Sheets: 3 (Product Spend Upload, Missing Categories, Vendor Breakdown)")
