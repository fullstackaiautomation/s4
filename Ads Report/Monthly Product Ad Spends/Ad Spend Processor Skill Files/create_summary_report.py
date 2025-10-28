import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load configuration
with open('config.json', 'r') as f:
    config = json.load(f)

month = config['month']
output_dir = config['paths']['output_dir'].replace("{month}", month)

# Load the upload sheet
upload_df = pd.read_csv(os.path.join(output_dir, f"{month} Product Spend Upload.csv"))

# Clean up numeric columns
def clean_currency(val):
    if pd.isna(val) or val == '' or val == 'NaN':
        return 0
    if isinstance(val, str):
        return float(val.replace('$', '').replace(',', ''))
    return float(val)

def clean_numeric(val):
    if pd.isna(val) or val == '' or val == 'NaN':
        return 0
    if isinstance(val, str):
        val = val.replace('%', '').strip()
    return float(val) if val else 0

# Convert to numeric
upload_df['Ad Spend Numeric'] = upload_df['Ad Spend'].apply(clean_currency)
upload_df['Revenue Numeric'] = upload_df['Revenue'].apply(clean_currency)
upload_df['Clicks Numeric'] = upload_df['Clicks'].apply(clean_numeric)

# Calculate ROAS (Revenue / Ad Spend)
upload_df['ROAS'] = upload_df.apply(
    lambda row: row['Revenue Numeric'] / row['Ad Spend Numeric'] if row['Ad Spend Numeric'] > 0 else 0,
    axis=1
)

# Calculate CPC (Ad Spend / Clicks)
upload_df['CPC Calculated'] = upload_df.apply(
    lambda row: row['Ad Spend Numeric'] / row['Clicks Numeric'] if row['Clicks Numeric'] > 0 else 0,
    axis=1
)

print("Report Data Preparation")
print("=" * 80)

# 1. Top 20 products by ad spend
print("1. Top 20 Products by Ad Spend")
top_20_spend = upload_df.nlargest(20, 'Ad Spend Numeric')[['SKU', 'Title', 'Vendor', 'Ad Spend Numeric', 'Revenue Numeric', 'ROAS']].copy()
top_20_spend['ROAS'] = top_20_spend['ROAS'].round(2)
print(f"   Found {len(top_20_spend)} products")

# 2. Top 20 products by revenue
print("2. Top 20 Products by Revenue")
top_20_revenue = upload_df.nlargest(20, 'Revenue Numeric')[['SKU', 'Title', 'Vendor', 'Ad Spend Numeric', 'Revenue Numeric', 'ROAS']].copy()
top_20_revenue['ROAS'] = top_20_revenue['ROAS'].round(2)
print(f"   Found {len(top_20_revenue)} products")

# 3. Top 20 CPC costs by SKU
print("3. Top 20 Highest CPC by SKU")
top_20_cpc = upload_df[upload_df['CPC Calculated'] > 0].nlargest(20, 'CPC Calculated')[['SKU', 'Title', 'Vendor', 'CPC Calculated', 'Clicks Numeric']].copy()
top_20_cpc.columns = ['SKU', 'Title', 'Vendor', 'CPC', 'Clicks']
print(f"   Found {len(top_20_cpc)} products")

# 4. Top 20 vendors by ad spend
print("4. Top 20 Vendors by Ad Spend")
vendor_spend = upload_df.groupby('Vendor').agg({
    'Ad Spend Numeric': 'sum',
    'Revenue Numeric': 'sum'
}).reset_index()
vendor_spend['ROAS'] = (vendor_spend['Revenue Numeric'] / vendor_spend['Ad Spend Numeric']).round(2)
vendor_spend = vendor_spend.sort_values('Ad Spend Numeric', ascending=False).head(20)
print(f"   Found {len(vendor_spend)} vendors")

# 5. Top 20 product categories with vendor, ad spend, revenue, ROAS
print("5. Top 20 Product Categories with Vendor Details")
category_vendor = upload_df.groupby(['Product Category', 'Vendor']).agg({
    'Ad Spend Numeric': 'sum',
    'Revenue Numeric': 'sum'
}).reset_index()
category_vendor['ROAS'] = (category_vendor['Revenue Numeric'] / category_vendor['Ad Spend Numeric']).round(2)
category_vendor = category_vendor.sort_values('Ad Spend Numeric', ascending=False).head(20)
print(f"   Found {len(category_vendor)} category-vendor combinations")

# Create Excel workbook
output_file = os.path.join(output_dir, f"{month} Product Spend Report.xlsx")
wb = pd.ExcelFile(output_file)
existing_sheets = wb.sheet_names
print(f"\nExisting sheets: {existing_sheets}")

# Load existing workbook and add new sheet
from openpyxl import load_workbook
wb = load_workbook(output_file)

# Remove old Summary Report sheet if it exists
if 'Summary Report' in wb.sheetnames:
    del wb['Summary Report']

ws = wb.create_sheet("Summary Report", 0)  # Insert at beginning

# Define styles
title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=14)
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True, size=10)
data_font = Font(size=10)
center_align = Alignment(horizontal="center", vertical="center")
currency_format = '$#,##0.00'
decimal_format = '0.00'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

current_row = 1

# Title
cell = ws.cell(row=current_row, column=1, value="2025-10 Ad Spend Performance Report")
cell.font = title_font
cell.fill = title_fill
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(f'A{current_row}:F{current_row}')
current_row += 2

# ==================== TOP 20 PRODUCTS BY AD SPEND ====================
ws.cell(row=current_row, column=1, value="Top 20 Products by Ad Spend").font = section_font
ws.cell(row=current_row, column=1).fill = section_fill
ws.merge_cells(f'A{current_row}:F{current_row}')
current_row += 1

# Headers
headers = ['SKU', 'Title', 'Vendor', 'Ad Spend', 'Revenue', 'ROAS']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
current_row += 1

# Data
for idx, row in top_20_spend.iterrows():
    ws.cell(row=current_row, column=1, value=row['SKU']).border = thin_border
    ws.cell(row=current_row, column=2, value=row['Title']).border = thin_border
    ws.cell(row=current_row, column=3, value=row['Vendor']).border = thin_border

    cell = ws.cell(row=current_row, column=4, value=row['Ad Spend Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=5, value=row['Revenue Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=6, value=row['ROAS'])
    cell.number_format = decimal_format
    cell.border = thin_border

    current_row += 1

current_row += 1

# ==================== TOP 20 PRODUCTS BY REVENUE ====================
ws.cell(row=current_row, column=1, value="Top 20 Products by Revenue").font = section_font
ws.cell(row=current_row, column=1).fill = section_fill
ws.merge_cells(f'A{current_row}:F{current_row}')
current_row += 1

# Headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
current_row += 1

# Data
for idx, row in top_20_revenue.iterrows():
    ws.cell(row=current_row, column=1, value=row['SKU']).border = thin_border
    ws.cell(row=current_row, column=2, value=row['Title']).border = thin_border
    ws.cell(row=current_row, column=3, value=row['Vendor']).border = thin_border

    cell = ws.cell(row=current_row, column=4, value=row['Ad Spend Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=5, value=row['Revenue Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=6, value=row['ROAS'])
    cell.number_format = decimal_format
    cell.border = thin_border

    current_row += 1

current_row += 1

# ==================== TOP 20 HIGHEST CPC ====================
ws.cell(row=current_row, column=1, value="Top 20 Highest CPC Costs by SKU").font = section_font
ws.cell(row=current_row, column=1).fill = section_fill
ws.merge_cells(f'A{current_row}:E{current_row}')
current_row += 1

# Headers
cpc_headers = ['SKU', 'Title', 'Vendor', 'CPC', 'Clicks']
for col_idx, header in enumerate(cpc_headers, 1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
current_row += 1

# Data
for idx, row in top_20_cpc.iterrows():
    ws.cell(row=current_row, column=1, value=row['SKU']).border = thin_border
    ws.cell(row=current_row, column=2, value=row['Title']).border = thin_border
    ws.cell(row=current_row, column=3, value=row['Vendor']).border = thin_border

    cell = ws.cell(row=current_row, column=4, value=row['CPC'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=5, value=row['Clicks'])
    cell.number_format = '0'
    cell.border = thin_border

    current_row += 1

current_row += 1

# ==================== TOP 20 VENDORS BY AD SPEND ====================
ws.cell(row=current_row, column=1, value="Top 20 Vendors by Ad Spend").font = section_font
ws.cell(row=current_row, column=1).fill = section_fill
ws.merge_cells(f'A{current_row}:D{current_row}')
current_row += 1

# Headers
vendor_headers = ['Vendor', 'Ad Spend', 'Revenue', 'ROAS']
for col_idx, header in enumerate(vendor_headers, 1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
current_row += 1

# Data
for idx, row in vendor_spend.iterrows():
    ws.cell(row=current_row, column=1, value=row['Vendor']).border = thin_border

    cell = ws.cell(row=current_row, column=2, value=row['Ad Spend Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=3, value=row['Revenue Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=4, value=row['ROAS'])
    cell.number_format = decimal_format
    cell.border = thin_border

    current_row += 1

current_row += 1

# ==================== TOP 20 CATEGORIES WITH VENDOR DETAILS ====================
ws.cell(row=current_row, column=1, value="Top 20 Product Categories with Vendor Details").font = section_font
ws.cell(row=current_row, column=1).fill = section_fill
ws.merge_cells(f'A{current_row}:E{current_row}')
current_row += 1

# Headers
category_headers = ['Category', 'Vendor', 'Ad Spend', 'Revenue', 'ROAS']
for col_idx, header in enumerate(category_headers, 1):
    cell = ws.cell(row=current_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center_align
current_row += 1

# Data
for idx, row in category_vendor.iterrows():
    ws.cell(row=current_row, column=1, value=row['Product Category']).border = thin_border
    ws.cell(row=current_row, column=2, value=row['Vendor']).border = thin_border

    cell = ws.cell(row=current_row, column=3, value=row['Ad Spend Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=4, value=row['Revenue Numeric'])
    cell.number_format = currency_format
    cell.border = thin_border

    cell = ws.cell(row=current_row, column=5, value=row['ROAS'])
    cell.number_format = decimal_format
    cell.border = thin_border

    current_row += 1

# Set column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12

# Save the workbook
wb.save(output_file)

print(f"\nCreated: {output_file}")
print(f"Summary Report sheet added with all 5 analysis sections")
print("\nReport Contents:")
print(f"  - Top 20 Products by Ad Spend")
print(f"  - Top 20 Products by Revenue")
print(f"  - Top 20 Highest CPC by SKU")
print(f"  - Top 20 Vendors by Ad Spend")
print(f"  - Top 20 Categories with Vendor Details")
